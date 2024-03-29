"""
Author:
    Pedro Antônio de Souza 201810557
"""


from openpyxl import load_workbook  # Importação biblioteca para leitura de arquivos .xlsx
from openpyxl import Workbook  # Importação para escrita de arquivos .xlsx
from collections import Counter  # Importação para comparação de valores em dicionários
from os import path  # Biblioteca para leitura de arquivos
import time  # Biblioteca para analisar o tempo de execução do algoritmo
import argparse


class Restricao(object):
    """
    Classe para instanciar objetos que contenham algum tipo de restrição, como professores e turmas.

    Attributes:
        nome (str): Nome do objeto que possui restrição.
        restricoes (list of Horario): Lista de horários restritos.
    """

    def __init__(self, nome):
        """
        Construtor da classe Restricao.

        Args:
            nome (str): Nome do objeto que possui restrição.
        """
        self.nome = nome
        self.restricoes = []

    def add_restricao(self, horario):
        """
        Metodo para adicionar um horário como restrição.

        Args:
            horario (Horario): Objeto do tipo Horario que será adicionado a lista de restrições.
        """
        self.restricoes.append(horario)


class Professor(Restricao):
    """
    Classe para instanciar objetos do tipo Professor. Essa classe herda de Restricao. Além de restrições, um objeto do tipo professor também possui preferências.

    Attributes:
        preferencias (list of Horario): Lista de horários preferidos.
        qtd_preferencias_atendidas (int): Quantidade de preferências atendidas.
    """

    def __init__(self, nome):
        """
        Construtor da classe Professor.

        Args:
            nome (str): Nome do professor.
        """

        super().__init__(nome)
        self.preferencias = []
        self.qtd_preferencias_atendidas = 0

    def add_preferencia(self, horario):
        """
        Metodo para adicionar um horário como preferência.

        Args:
            horario (Horario): Objeto do tipo Horario que será adicionado a lista de preferências.
        """
        self.preferencias.append(horario)

    def tem_preferencia(self, horario):
        """
        Metodo para verificar se professor tem preferência por determinado horário.

        Args:
            horario (Horario): Objeto do tipo Horario.

        Returns:
            bool: True se o horário for preferência e False caso contrário.
        """
        return horario in self.preferencias

    def qtd_preferencias_nao_atendidas(self):
        """
        Retorna a quantidade de preferências não atendidas.

        Returns:
            (int|None): Quantidade de preferências não atendidas ou None caso o professor não possua preferências.
        """
        return None if len(self.preferencias) == 0 else len(self.preferencias) - self.qtd_preferencias_atendidas

    def preferencia_atendida(self):
        """
        Incrementa em um a quantidade de preferências atendidas do professor.
        """
        self.qtd_preferencias_atendidas += 1


class Vertice(object):
    """
    Classe para instanciar vértices do grafo. São as aulas da instituição.

    Attributes:
        materia (string): Matéria (disciplina) da aula.
        turma (Restricao): Turma da aula.
        professor (Professor): Professor que leciona a matéria para a turma.
        grau_saturacao (int): Grau de saturação do vértice. Quantos vizinhos coloridos ele possui.
        restricoes (list of Horario): Lista de horários restritos.
        restricoes_leves (list of Horario): Lista de horários com restrições leves. Restrições que podem ser ignoradas, caso necessário.
        preferencias (list of Horario): Lista de horários preferidos.
        horarios_sequencia (list of Horario): Lista de horários em que a mesma aula é lecionada.
   """

    def __init__(self, materia, turma, professor):
        """
        Construtor da classe Vertice.

        Args:
            materia (string): Matéria (disciplina) da aula.
            turma (Restricao): Turma da aula.
            professor (Professor): Professor que leciona a matéria para a turma.
       """
        self.materia = materia
        self.turma = turma
        self.professor = professor
        self.grau_saturacao = 0
        self.restricoes = []
        self.restricoes_leves = []  # 3 aulas seguidas da mesma matéria
        self.preferencias = []
        self.horarios_sequencia = []

    def tem_restricao(self, horario, leves=False):
        """
        Método para verificar se determinado horário é restrito.

        Args:
            horario (Horario): Horário que deve-se verificar a restrição.
            leves (bool): True para verificar também as restrições leves, False para não verificá-las.

        Returns:
            bool: True caso haja restrição, False para o oposto.
        """
        if leves:
            return horario in (self.restricoes + self.restricoes_leves)
        return horario in self.restricoes

    def qtd_restricao(self, leves=False):
        """
        Método para verificar a quantidade de restrições existentes.

        Args:
            leves (bool): True para contabilizar também as restrições leves, False para não contabilizá-las.

        Returns:
            int: Quantidade de restrições que o vértice possui.
        """
        if not leves:
            return len(self.restricoes)
        return len(self.restricoes) + len(self.restricoes_leves)

    def add_restricao(self, horario):
        """
        Adiciona um horário às restrições do vértice.

        Além disso, remove o horário de self.preferencias e self.horarios_sequencia, caso seja um item dessas listas.

        Args:
            horario (Horario): Horário a ser adicionado como restrição.
        """

        if horario is not None and horario in self.preferencias:
            self.preferencias.remove(horario)

        if horario is not None and horario in self.horarios_sequencia:
            self.horarios_sequencia.remove(horario)

        if not self.tem_restricao(horario):
            self.restricoes.append(horario)

    def atualiza_restricoes(self):
        """
        Atualiza as restrições do vértice de acordo com as restrições do professor e da turma.

        Lê cada restrição e a adciona a lista de restrições do próprio vértice.
        """
        for horario in self.professor.restricoes:
            self.add_restricao(horario)

        for horario in self.turma.restricoes:
            self.add_restricao(horario)

    def tem_restricao_leve(self, horario):
        """
        Metodo para verificar se o vértice possui restrição leve em determinado horário.

        Args:
            horario (Horario): Objeto do tipo Horario.

        Returns:
            bool: True se o horário for restrição leve e False caso contrário.
        """
        return horario in self.restricoes_leves

    def add_restricao_leve(self, horario):
        """
        Adiciona um horário às restrições leves do vértice.

        Além disso, remove o horário de self.preferencias e self.horarios_sequencia, caso seja um item dessas listas.

        Args:
            horario (Horario): Horário a ser adicionado como restrição leve.
        """

        if horario is not None and horario in self.preferencias:
            self.preferencias.remove(horario)

        if horario is not None and horario in self.horarios_sequencia:
            self.horarios_sequencia.remove(horario)

        if horario is not None and not self.tem_restricao_leve(horario):
            self.restricoes_leves.append(horario)

    def qtd_preferencia(self):
        """
        Método para verificar a quantidade de preferências do vértice.

        Returns:
            int: Quantidade de preferências que o vértice possui.
        """
        return len(self.preferencias)

    def add_preferencia(self, horario):
        """
        Adiciona um horário às preferências do vértice, caso o horário não seja uma restrição.

        Args:
            horario (Horario): Horário a ser adicionado como preferência.
        """
        if (horario is not None
                and not self.tem_restricao(horario, True)):
            self.preferencias.append(horario)

    def atualiza_preferencias(self):
        """
        Atualiza as preferências do vértice de acordo com as preferências do professor.

        Lê cada preferência e a adciona a lista de preferências do próprio vértice.
        """
        for horario in self.professor.preferencias:
            self.add_preferencia(horario)

    def add_horario_sequencia(self, horario):
        """
        Adiciona horário que pode ser atribuido ao vértice para que sejam ministradas aulas iguais em horários sequenciais.

        Args:
            horario (Horario): Horário a ser adicionado como horário em sequencia.
        """
        if horario not in self.horarios_sequencia:
            self.horarios_sequencia.append(horario)

    def tem_horario_sequencia(self):
        """
        Verifica se já há sugestão de horário em sequência.

        Returns:
            int: Quantidade de horários sugeridos para que a aula seja ministrada em horário sequencial a outras aulas iguais.
        """
        return len(self.horarios_sequencia)

    def eh_horario_sequencia(self, horario):
        """
        Verifica se determinado horário é sugerido ao vértice como um horário sequencial a outra aula igual.

        Args:
            horario (Horario): Horário que se deseja verificar.

        Returns:
             bool: True se horario é um horário sugerido, False caso contrário.
        """
        return horario in self.horarios_sequencia

    def eh_igual(self, outro_vertice):
        """
        Verifica se o vertice é igual a outro vértice.

        Verifica-se a igualdade em relação a matéria, turma e professor.

        Args:
            outro_vertice (Vertice): Vertice para comparação.

        Returns:
            bool: True caso sejam iguais, False caso contrário.

        """

        return (outro_vertice.materia == self.materia
                and outro_vertice.turma == self.turma
                and outro_vertice.professor == self.professor)

    def deve_ser_vizinho(self, outro_vertice):
        """
        Verifica se o vértice deve ser vizinho de um outro vertice.

        A verificação é feita analisando se os vértices são iguais completamente,
        ou se apenas as turmas são iguais ou se apenas os professores são iguais

        Args:
            outro_vertice (Vertice): Vertice para análise.

        Returns:
            bool: True caso devam ser vizinhos, False caso contrário.

        """

        if (self.turma == outro_vertice.turma
                or self.professor == outro_vertice.professor
                or self.eh_igual(outro_vertice)):
            return True
        return False

    def incrementa_grau_saturacao(self):
        """
        Incrementa o grau de saturação do vértice em um.
        """
        self.grau_saturacao += 1

    def horarios_preferidos(self):
        '''
        Retorna lista de horários preferidos.

        Return:
            list of Horario: Lista de horários por ordem de frequência na lista de preferências concatenada com a lista de horários em sequência.
        '''
        return [horario[0] for horario in Counter(self.preferencias + self.horarios_sequencia).most_common()]

    def dados(self):
        return 'Mat: {}, Tur: {}, Pro: {}'.format(self.materia, self.turma.nome, self.professor.nome)


class Horario(object):
    """
    Classe que define horários que as aulas podem ser ministradas

    Attributes:
        dia (str): Dia da semana.
        hora (str): Hora.
        vertices (list of Vertice): Aulas ministradas nesse horário.
    """

    def __init__(self, dia, hora):
        """
        Construtor da classe Horario.

        Args:
            dia (str): Dia da semana.
            hora (str): Hora.
            vertices (list of Vertice): Aulas ministradas nesse horário.
        """

        self.dia = dia
        self.hora = hora
        self.vertices = []

    def add_vertice(self, vertice):
        """
        Adiciona vértices (aulas) ao horário.

        Args:
            vertice (Vertice): Aula alocada para esse horário.
        """
        igual = False
        for colorido in self.vertices:
            igual = vertice.eh_igual(colorido)

        self.vertices.append(vertice)

    def turma_tem_aula(self, turma):
        for vertice in self.vertices:
            if vertice.turma.nome == turma:
                return [vertice.materia, vertice.professor.nome]
        return False

    def dados(self):
        return '{} {}'.format(self.dia, self.hora)


class HorarioDeAulas(object):
    """
    Classe com métodos para ler uma planilha, analisar os dados e alocar as aulas nos horários permitidos.

    Attributes:
        arquivo (str): Nome do arquivo passado como parâmetro.
        planilha (Workbook): Arquivo .xlsx lido com a biblioteca openpyxl
        turmas (dict of str: Restricao): Dicionário indexado pelo nome da turma e objeto do tipo Restricao como valor.
        professores (dict of str: Professor): Dicionário indexado pelo nome do professor e objeto do tipo Professor como valor.
        horarios (dict of str: (dict of str: Horario)): Dicionário indexado na primeira camada com os nomes dos dias e na segunda com horários.
        aulas_por_dia (int): Número de aulas que a instituição oferece por dia.
        vertice (list of Vertice): Lista de vértices (aulas) do grafo.
        quantidade_vertices_sem_horario (int): Quantidade de vertices sem horário definido.
        lista_adjacencia (dict of Vertice: (list of Vertice)): Dicionário indexado por objetos Vertice que possuem uma lista de objetos Vertice adjacentes.
        total_preferencias (int): Total de preferências dos professores
        preferencias_atendidas (int): Total de preferências atendidas após aplicação do algoritmo de coloração
        prioridade_aula_sequencial (bool): True para priorizar o máximo de aulas em sequencia igual a 2, False para não analisar isso. False é o valor default.
    """

    def __init__(self, arquivo, prioridade_aula_sequencial=False):
        """
        Construtor da classe HorarioDeAulas.

        Args:
            arquivo (str): Caminho do arquivo que se deseja ler.
            prioridade_aula_sequencial (bool): True para priorizar o máximo de aulas em sequencia igual a 2, False para não analisar isso. False é o valor default.
        """
        self.arquivo = arquivo
        self.planilha = load_workbook(arquivo)
        self.turmas = dict()
        self.professores = dict()
        self.horarios = dict()
        self.aulas_por_dia = 0
        self.vertices = list()
        self.quantidade_vertices_sem_horario = 0
        self.lista_adjacencia = dict()
        self.total_preferencias = 0
        self.preferencias_atendidas = 0
        self.prioridade_aula_sequencial = prioridade_aula_sequencial

        self.inicializa_vertices()
        self.inicializa_horarios()
        self.define_restricoes_professores()
        self.define_restricoes_turmas()
        self.define_preferencias_professores()
        self.atualiza_restricoes_preferencias_vertices()
        self.define_arestas()

    def inicializa_vertices(self):
        """
        Lê a planilha Dados e cria vértices a partir desses dados.

        Além disso, atribui valores às listas de professores e turmas.
        """
        informacoes = self.planilha['Dados']

        if informacoes.max_row == 1:
            # Verifica se a planilha possui apenas o cabeçalho.
            # Em caso verdadeiro, abandona aqui a execução do método.
            return

        informacoes.delete_rows(1)  # Remove cabecalho

        # Laço para criar vertices com a tupla (Matéria, Turma, Professor)
        for linha in informacoes:
            materia = linha[0].value
            turma = linha[1].value
            professor = linha[2].value
            quantidade = linha[3].value

            # Verifica se a turma já está no dicionario de turmas
            # Se nao estiver, cria uma nova turma do tipo Restricao e adiciona ao dicionario
            if turma in self.turmas:
                turma = self.turmas[turma]
            else:
                turma = Restricao(turma)
                self.turmas[turma.nome] = turma

            # Verifica se o professor já está no dicionario de professores
            # Se nao estiver, cria um novo professor do tipo Professor e adiciona ao dicionario
            if professor in self.professores:
                professor = self.professores[professor]
            else:
                professor = Professor(professor)
                self.professores[professor.nome] = professor

            # Insere os vertices
            self.insere_vertice((materia, turma, professor), quantidade)

    def inicializa_horarios(self):
        """
        Lê a planilha Configuracoes e instância objetos do tipo Horario.

        Além disso, atribui valores ao dicionário de horários da classe.
        """
        informacoes = self.planilha['Configuracoes']

        if informacoes.max_row == 1:
            # Verifica se a planilha possui apenas o cabeçalho.
            # Em caso verdadeiro, abandona aqui a execução do método.
            return

        informacoes.delete_rows(1)  # Remove cabecalho

        dias = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
        # Laço para inicializar horarios
        for dia in dias:
            self.horarios[dia] = dict()

            for hora in informacoes:
                horario = Horario(dia, str(hora[0].value))
                self.horarios[dia][str(hora[0].value)] = horario

            self.aulas_por_dia = len(self.horarios[dia])

    def define_restricoes_professores(self):
        """
        Lê a planilha Restricao e adiciona restrições aos objetos do tipo Professor.
        """
        informacoes = self.planilha['Restricao']

        if informacoes.max_row == 1:
            # Verifica se a planilha possui apenas o cabeçalho.
            # Em caso verdadeiro, abandona aqui a execução do método.
            return

        informacoes.delete_rows(1)  # Remove cabecalho

        for restricao in informacoes:
            professor = restricao[0].value
            hora = str(restricao[1].value)
            dia = restricao[2].value

            if (professor in self.professores
                    and dia in self.horarios
                    and hora in list(self.horarios.values())[0]):
                # Verifica se o professor foi alocado em algum vértice,
                # se o dia, é um dia letivo
                # e se o horário é um horário de inicio de aula.
                horario = self.horarios[dia][hora]
                self.professores[professor].add_restricao(horario)

    def define_restricoes_turmas(self):
        """
        Lê a planilha Restricao Turma e adiciona restrições aos objetos do tipo Turma.
        """
        informacoes = self.planilha['Restricoes Turma']

        if informacoes.max_row == 1:
            # Verifica se a planilha possui apenas o cabeçalho.
            # Em caso verdadeiro, abandona aqui a execução do método.
            return

        informacoes.delete_rows(1)  # Remove cabecalho

        for restricao in informacoes:
            turma = restricao[0].value
            hora = str(restricao[1].value)
            dia = restricao[2].value

            if (turma in self.turmas
                    and dia in self.horarios
                    and hora in list(self.horarios.values())[0]):
                # Verifica se a turma exite na lista de turmas,
                # se o dia é um dia letivo
                # e se o horário é um horário de início de aula.
                horario = self.horarios[dia][hora]
                self.turmas[turma].add_restricao(horario)

    def define_preferencias_professores(self):
        """
        Lê a planilha Preferencias e adiciona as preferências aos objetos do tipo Professor.
        """
        informacoes = self.planilha['Preferencias']

        if informacoes.max_row == 1:
            # Verifica se a planilha possui apenas o cabeçalho.
            # Em caso verdadeiro, abandona aqui a execução do método.
            return

        informacoes.delete_rows(1)  # Remove cabecalho

        for preferencia in informacoes:
            professor = preferencia[0].value
            hora = str(preferencia[1].value)
            dia = preferencia[2].value

            if (professor in self.professores
                    and dia in self.horarios
                    and hora in list(self.horarios.values())[0]):
                # Verifica se o professor foi alocado em algum vértice,
                # se o dia, é um dia letivo
                # e se o horário é um horário de inicio de aula.
                horario = self.horarios[dia][hora]
                self.professores[professor].add_preferencia(horario)
                self.total_preferencias += 1

    def atualiza_restricoes_preferencias_vertices(self):
        """
        Para cada vértice, chama os respectivos métodos para atualizar as restrições e as preferências dos vértices.
        """
        for vertice in self.vertices:
            vertice.atualiza_restricoes()
            vertice.atualiza_preferencias()

    def define_arestas(self):
        """
        Cria as arestas do grafo.

        Percorre cada vértice. Ao percorrer, verifica quais vertices devem ser vizinhos através do metodo deve_ser_vizinho().

        """
        for indice_v1 in range(len(self.vertices)):
            vertice_1 = self.vertices[indice_v1]
            self.lista_adjacencia[vertice_1] = []

            for indice_v2 in range(len(self.vertices)):
                if indice_v1 != indice_v2:
                    vertice_2 = self.vertices[indice_v2]

                    if vertice_1.deve_ser_vizinho(vertice_2):
                        # Se o vértice deve ser vizinho, o adiciona à lista de adjacência.
                        self.lista_adjacencia[vertice_1].append(vertice_2)

    def insere_vertice(self, dados, qtd=1):
        """
        Cria objetos do tipo Vertice e adiciona-os à lista de vértices.

        Args:
            dados (tuple of (str, Restricao, Professor)): Dados do vértice (aula).
            qtd (int): Quantidade que aulas desse tipo devem ser ministradas.
        """
        materia = dados[0]
        turma = dados[1]
        professor = dados[2]
        for i in range(qtd):
            self.quantidade_vertices_sem_horario += 1
            self.vertices.append(Vertice(materia, turma, professor))

    '''
    Aqui se inicia os algoritmos para coloração.

    Foi utilizado o algoritmo de DSATUR com algumas adaptações.
    No algoritmo original, prioriza-se apenas os vértices com maior grau de saturação.
    Nessa heurística são priorizados em ordem:
        - Vértices com maior número de restrições;
        - Vértices com maior grau de saturação;
        - Vértices com maior grau no grafo original;
        - Vértices com possibilidade de horário em sequência;
        - Vértices com mais preferências.
    '''

    def dsatur_com_heristica(self):
        """
        Método para atribuição dos horários à cada aula.

        Returns:
            list of Vertice: Lista de vértices com horário não definido.
        """
        lista_de_horarios = []  # Lista com objetos Horario (cores) possiveis.
        for dia in self.horarios:
            for horario in self.horarios[dia]:
                # Percorre cada objeto no dicionario de horários e os adiciona a lista_de_horarios
                lista_de_horarios.append(self.horarios[dia][horario])

        vertices_nao_coloridos = self.vertices.copy()
        vertices_nao_alocados = []

        # Inicia aqui a escolha do vertice a ser colorido
        # e o processo de coloração (atribuição dos horários)
        vertice_escolhido = self.escolher_vertice(vertices_nao_coloridos)
        vertices_nao_coloridos.remove(vertice_escolhido)
        horario = self.escolher_horario(vertice_escolhido, lista_de_horarios)
        self.define_horario(vertice_escolhido, horario, lista_de_horarios, vertices_nao_coloridos)

        while len(vertices_nao_coloridos) > 0:
            vertice_escolhido = self.escolher_vertice(
                self.vizinhos_nao_coloridos(vertice_escolhido, vertices_nao_coloridos))
            # Se não houver vizinhos sem cor, escolhe um vértice dentre todos os vértices não coloridos.
            vertice_escolhido = self.escolher_vertice(
                vertices_nao_coloridos) if vertice_escolhido is None else vertice_escolhido
            vertices_nao_coloridos.remove(vertice_escolhido)
            horario = self.escolher_horario(vertice_escolhido, lista_de_horarios)

            if horario is not None:
                # Se encontrar um horário em que o vértice se encaixa, define esse horário.
                if not vertice_escolhido.tem_restricao(horario):
                    self.define_horario(vertice_escolhido, horario, lista_de_horarios, vertices_nao_coloridos)
                else:
                    vertices_nao_alocados.append(vertice_escolhido)
            else:
                # Caso não encontre um horário para o vértice, o separa em uma lista.
                vertices_nao_alocados.append(vertice_escolhido)

        return vertices_nao_alocados

    def vertices_de_maior_grau(self, lista_de_vertices):
        """
        Encontra os vértices com maior grau no grafo original dentre uma lista de vertices.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para comparação.

        Returns:
            (list of Vertice): Lista com os vértices de maior grau.
        """
        vertices = []
        maior_grau = 0

        for vertice in lista_de_vertices:
            grau = len(self.lista_adjacencia[vertice])
            if grau > maior_grau:
                maior_grau = grau
                vertices = [vertice]
            elif grau == maior_grau:
                vertices.append(vertice)

        return vertices

    def vertices_maior_grau_saturacao(self, lista_de_vertices):
        """
        Encontra os vértices de maior grau de saturação dentre uma lista de vertices.

        Caso haja empate, encontra os vértices com maior grau no grafo original dentre os empatados.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para comparação.

        Returns:
            (list of Vertice): Lista com os vértices de maior grau de saturação.
        """
        vertices = []
        maior_grau = 0

        for vertice in lista_de_vertices:
            grau = vertice.grau_saturacao
            if grau > maior_grau:
                maior_grau = grau
                vertices = [vertice]
            elif grau == maior_grau:
                vertices.append(vertice)

        return self.vertices_de_maior_grau(vertices)

    def vertices_com_mais_restricoes(self, lista_de_vertices):
        """
        Encontra os vértices de maior quantidade de restrições dentre uma lista de vertices.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para comparação.

        Returns:
            (list of Vertice): Lista com os vértices de maior quantidade de restrições.
        """

        vertices = []
        maior_quantidade = 0

        for vertice in lista_de_vertices:
            quantidade = vertice.qtd_restricao()
            if quantidade > maior_quantidade:
                maior_quantidade = quantidade
                vertices = [vertice]
            elif quantidade == maior_quantidade:
                vertices.append(vertice)

        return vertices

    def vertices_com_horario_sequencia(self, lista_de_vertices):
        """
        Encontra os vértices que possuem sugestão de horário em sequencia dentre uma lista de vertices.

        Caso nenhum vértice possua a sugestão, a lista passada como argumento é retornada.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para comparação.

        Returns:
            (list of Vertice): Lista com os vértices que possuem sugestão de horário.
        """
        lista_aux = [vertice for vertice in lista_de_vertices if vertice.tem_horario_sequencia()]

        return lista_aux if len(lista_aux) > 0 else lista_de_vertices

    def vertices_com_mais_preferencias(self, lista_de_vertices):
        """
        Encontra os vértices com maior quantidade de preferências dentre uma lista de vertices.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para comparação.

        Returns:
            (list of Vertice): Lista com os vértices de maior quantidade de preferências.
        """
        vertices = []
        maior_quantidade = 0

        for vertice in lista_de_vertices:
            quantidade = vertice.qtd_preferencia()
            if quantidade > maior_quantidade:
                maior_quantidade = quantidade
                vertices = [vertice]
            elif quantidade == maior_quantidade:
                vertices.append(vertice)

        return vertices

    def escolher_vertice(self, lista_de_vertices):
        """
        Escolhe o melhor vértice para aplicação da coloração de acordo com a heurística.

        Aplica-se, em sequência que respeite os critérios da heurística, os métodos para filtrar os vértices da lista
        até que se encontre o melhor vértice.

        Args:
            lista_de_vertices (list of Vertice): Lista com os vértices para busca.

        Returns:
            (Vertice): Melhor vértice encontrado ou None caso não encontre nenhum.

        """
        lista_de_vertices = self.vertices_com_mais_restricoes(lista_de_vertices)
        lista_de_vertices = self.vertices_maior_grau_saturacao(lista_de_vertices)
        lista_de_vertices = self.vertices_com_horario_sequencia(lista_de_vertices)
        lista_de_vertices = self.vertices_com_mais_preferencias(lista_de_vertices)

        return None if len(lista_de_vertices) == 0 else lista_de_vertices[0]

    def escolher_horario(self, vertice, lista_de_horarios):
        """
        Escolhe o melhor horário (cor) para um determinado vértice dada uma lista de horários.

        Args:
            vertice (Vertice): Vértice que se deseja definir horário (cor).
            lista_de_horarios (list of Horario): Lista de horários (cores) disponíveis.

        Returns:
            (Horario|None): Melhor horário encontrado ou None caso não encontre nenhum horário.
        """
        horarios_preferidos = vertice.horarios_preferidos()

        for horario in horarios_preferidos:
            # Percorre cada horário preferido verificando se algum é horário restrito para o vértice.
            if vertice.tem_restricao(horario):
                # Caso seja, o retira da lista.
                horarios_preferidos.remove(horario)

        if len(horarios_preferidos) > 0:
            return horarios_preferidos[0]
        else:
            # Se não houver horários preferidos, deve-se buscar dentre todos os horários disponíveis iniciando do menor
            # horário utilizado até o momento.
            horarios_com_restricao_leve = []
            id_horario = self.indice_menor_horario_utilizado(lista_de_horarios)
            flag = 0
            qtd_horarios = len(lista_de_horarios)
            encontrado = False
            while flag < qtd_horarios and not encontrado:
                horario = lista_de_horarios[id_horario % qtd_horarios]

                if vertice.tem_restricao_leve(horario) and not vertice.tem_restricao(horario):
                    # Se o vértice possui apenas restrição leve (restrição de 3 aulas em sequência), adiciona horário na
                    # lista de restrições leves
                    horarios_com_restricao_leve.append(horario)

                if not vertice.tem_restricao(horario, True):
                    encontrado = True
                else:
                    id_horario += 1
                flag += 1

            if encontrado:
                return lista_de_horarios[id_horario % qtd_horarios]
            elif not encontrado and len(horarios_com_restricao_leve) > 0:
                return horarios_com_restricao_leve[0]
            else:
                return None

    def define_horario(self, vertice, horario, lista_de_horarios, vertices_nao_coloridos):
        """
        Define horário (colore) de um determinado vértice.

        Args:
            vertice (Vertice): Vértice que será definido um horário (cor).
            horario (Horario): Horário (cor) definido para o vértice.
            lista_de_horarios (list of Horario): Lista de horários disponíveis.
            vertices_nao_coloridos (list of Vertice): Lista de vértices não coloridos)
        """
        horario.add_vertice(vertice)
        self.quantidade_vertices_sem_horario -= 1
        # Verifica se horário é preferência do professor e incrementa em um a quantidade de preferências atendidas.
        if vertice.professor.tem_preferencia(horario):
            vertice.professor.preferencia_atendida()
            self.preferencias_atendidas = self.preferencias_atendidas + 1

        horario_seguinte = lista_de_horarios.index(horario) + 1
        horario_anterior = lista_de_horarios.index(horario) - 1

        # Não faz sentido dar preferencia em aula em sequencia, caso ela seja a primeira do dia seguinte por isso a
        # condicao após o and
        horario_seguinte = (horario_seguinte if (horario_seguinte < len(lista_de_horarios)
                                                 and horario_seguinte % self.aulas_por_dia != 0)
                            else None)

        # Não faz sentido dar preferencia em aula em sequencia, caso ela seja a última do dia anterior por isso a
        # condicao após o and
        horario_anterior = (horario_anterior if (horario_anterior >= 0
                                                 and horario_anterior % self.aulas_por_dia != self.aulas_por_dia - 1)
                            else None)

        for vizinho in self.vizinhos_nao_coloridos(vertice, vertices_nao_coloridos):
            vizinho.incrementa_grau_saturacao()
            # Adiciona o horário que está sendo colorido como restrição a todos os vértices vizinhos.
            vizinho.add_restricao(horario)
            if self.prioridade_aula_sequencial and vizinho.eh_igual(vertice):
                if vertice.eh_horario_sequencia(horario):
                    # Se o vértice que está sendo definido o horario já for uma aula em sequência, adiciona o horário seguinte e
                    # anterior como restrição leve aos vértices vizinhos que são a mesma materia e turma.
                    if horario_seguinte is not None:
                        vizinho.add_restricao_leve(lista_de_horarios[horario_seguinte])

                    if horario_anterior is not None:
                        vizinho.add_restricao_leve(lista_de_horarios[horario_anterior])
                else:
                    # Se não for aula em sequência, adiciona o horário seguinte e anterior como preferencia aos vértices iguais.
                    if horario_seguinte is not None:
                        vizinho.add_horario_sequencia(lista_de_horarios[horario_seguinte])

                    if horario_anterior is not None:
                        vizinho.add_horario_sequencia(lista_de_horarios[horario_anterior])

    def vizinhos_nao_coloridos(self, vertice, vertices_nao_coloridos):
        """
        Encontra todos os vizinhos de um determinado vértice que ainda não estão coloridos.

        Args:
            vertice (Vertice): Vértice que se deseja encontrar a vizinha descolorida.
            vertices_nao_coloridos (list of Vertice): Lista de vértices não coloridos.

        Return:
            (list of Vertice): Vizinhos não coloridos.
        """
        return [vizinho for vizinho in self.lista_adjacencia[vertice] if vizinho in vertices_nao_coloridos]

    def indice_menor_horario_utilizado(self, lista_de_horarios):
        """
        Encontra o horário de menor índice utilizado em uma determinada lista de horários.

        Args:
            lista_de_horarios (list of Horario): Lista de horários.

        Returns:
            (int): Menor índice como horário utilizado.
        """
        indice = 0;
        encontrado = False
        while not encontrado and indice < len(lista_de_horarios):
            if len(lista_de_horarios[indice].vertices) > 0:
                encontrado = True
            else:
                indice += 1

        return indice

    def proporcao_preferencias_atendidas(self):
        """
        Calcula a proporção entre a quantidade de preferências atendidas e o total de preferências existentes.

        Returns:
            (float): Proporção entre quantidade de preferências atendidas e o total de preferências existentes.
        """
        return self.preferencias_atendidas / self.total_preferencias

    def quantidade_horarios_utilizados(self):
        """
        Encontra a quantidade de horários (cores) utilizados.

        Returns:
            (int): Quantidade de horários utilizados.
        """
        qtd = 0
        for dia in self.horarios:
            for horario in self.horarios[dia]:
                if len(self.horarios[dia][horario].vertices) > 0:
                    qtd += 1
        return qtd

    def gerar_horarios_por_turma(self, nome_arquivo_saida):
        """
        Gera um arquivo .xlsx com os horários das turmas.

        Args:
            nome_arquivo_saida (list of Horario): Nome do arquivo de saída.
        """
        dias = list(self.horarios.keys())
        horas = list(self.horarios[dias[0]].keys())

        lista_de_horarios = []
        for dia in dias:
            for horario in horas:
                lista_de_horarios.append(self.horarios[dia][horario])

        planilha = Workbook()
        for turma in self.turmas:
            horario = planilha.create_sheet(str(turma))
            horario.cell(row=1, column=1, value=str(turma))

            coluna = 2
            for dia in dias:
                horario.cell(row=2, column=coluna, value=dia)
                coluna += 1
                horario.column_dimensions[chr(coluna + 95)].width = 20

            linha = 3
            for hora in horas:
                horario.cell(row=linha, column=1, value=hora)
                linha += 1

        for horario in lista_de_horarios:
            coluna = dias.index(horario.dia) + 2
            linha = horas.index(horario.hora) + 3
            for vertice in horario.vertices:
                materia = vertice.materia
                turma = str(vertice.turma.nome)
                professor = vertice.professor.nome

                planilha[turma].cell(row=linha, column=coluna, value='{} ({})'.format(materia, professor))

        del planilha['Sheet']
        planilha.save(nome_arquivo_saida + '.xlsx')

    def gerar_horarios_por_professor(self, nome_arquivo_saida):
        """
        Gera um arquivo .xlsx com os horários dos professores.

        Args:
            nome_arquivo_saida (list of Horario): Nome do arquivo de saída.
        """
        dias = list(self.horarios.keys())
        horas = list(self.horarios[dias[0]].keys())

        lista_de_horarios = []
        for dia in dias:
            for horario in horas:
                lista_de_horarios.append(self.horarios[dia][horario])

        planilha = Workbook()
        for professor in sorted(self.professores):
            horario = planilha.create_sheet(str(professor))
            horario.cell(row=1, column=1, value=str(professor))

            coluna = 2
            for dia in dias:
                horario.cell(row=2, column=coluna, value=dia)
                coluna += 1
                horario.column_dimensions[chr(coluna + 95)].width = 20

            linha = 3
            for hora in horas:
                horario.cell(row=linha, column=1, value=hora)
                linha += 1

        for horario in lista_de_horarios:
            coluna = dias.index(horario.dia) + 2
            linha = horas.index(horario.hora) + 3
            for vertice in horario.vertices:
                materia = vertice.materia
                turma = str(vertice.turma.nome)
                professor = vertice.professor.nome

                planilha[professor].cell(row=linha, column=coluna, value='{} ({})'.format(turma, materia))

        del planilha['Sheet']
        planilha.save(nome_arquivo_saida + '.xlsx')

    def imprimir_resultados(self):
        """
        Método para impressão dos resultados obtidos.
        """
        nome_escola = self.arquivo.split('/')[-1].split('.xlsx')[0]
        print(nome_escola + ':')
        print('\nQuantidade de cores:', self.quantidade_horarios_utilizados())
        print('\nVértices não coloridos:', self.quantidade_vertices_sem_horario)
        print('\nPreferências atendidas sobre o total de preferências:', self.proporcao_preferencias_atendidas())
        print('\nQuantidade de preferências não atendidas para cada professor (somente dos professores que possuem preferências):')
        for professor in sorted(self.professores):
            professor = self.professores[professor]
            if professor.qtd_preferencias_nao_atendidas() is not None:
                print('{}: {}'.format(professor.nome, professor.qtd_preferencias_nao_atendidas()))


def main():
    argumentos = argparse.ArgumentParser()
    argumentos.add_argument('--file', action='store', dest='arquivo',
                        default='', required=True,
                        help='Arquivo de extensão .xlxs com os dados da escola.')

    argumentos.add_argument('--aulas-sequenciais', action='store', dest='aulas_sequenciais',
                        default='N', choices=['S', 'N'], required=False,
                        help='[S/N] Opção de priorizar aulas sequencias.')

    argumentos.add_argument('--gerar-horarios-turmas', action='store', dest='horario_turma',
                        default=False, required=False,
                        help='Nome do desejado para o arquivo gerado com os horários de aula por turmas.')

    argumentos.add_argument('--gerar-horarios-professores', action='store', dest='horario_professor',
                        default=False, required=False,
                        help='Nome do desejado para o arquivo gerado com os horários de aula por professores.')

    args = argumentos.parse_args()
    arquivo = args.arquivo
    aulas_sequenciais = True if args.aulas_sequenciais == 'S' else False
    horario_turma = args.horario_turma
    horario_professor = args.horario_professor

    if not path.exists(arquivo):
        print("Arquivo não encontrado.")
    else:
        horarios_de_aula = HorarioDeAulas(arquivo, prioridade_aula_sequencial=aulas_sequenciais)
        tempo_1 = time.time()
        vertices_sem_horario = horarios_de_aula.dsatur_com_heristica()
        tempo_2 = time.time()

        if(horario_turma):
            horarios_de_aula.gerar_horarios_por_turma(horario_turma)

        if(horario_professor):
            horarios_de_aula.gerar_horarios_por_professor(horario_professor)

        horarios_de_aula.imprimir_resultados()
        print('Tempo de execução:', tempo_2-tempo_1)

if __name__ == '__main__':
    main()
